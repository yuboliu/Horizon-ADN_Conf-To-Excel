import pprint
import re
import os
import openpyxl
from collections import defaultdict

file_path = "startup-config"  # Replace with your actual file path
pattern_ipnatpool_matches = re.compile(r'^ip nat pool .*')
pattern_healthCheck_matches = re.compile(r'^health check .*')
pattern_slbPool_matches = re.compile(r'^slb pool .*')
pattern_slbProcotolProfile_matches = re.compile(r'^slb profile\b(?!.*(?:node|persist)).*$')
pattern_slbPresistProfile_matches = re.compile(r'^slb profile persist .*')
pattern_slbVA_matches = re.compile(r'^slb virtual-address .*')
pattern_list = (pattern_ipnatpool_matches, pattern_healthCheck_matches, pattern_slbPool_matches, pattern_slbVA_matches, pattern_slbPresistProfile_matches, pattern_slbProcotolProfile_matches)
pattern_va_port_matches = re.compile(r'port\b\s\d+.*')
pattern_hex_unicode = re.compile(r'%u[0-9A-Fa-f]{4}')
pattern_hex_split_unicode = re.compile(r'(%u[0-9A-Fa-f]{4})')
# vaPortDict = defaultdict(dict)
outputDict = defaultdict(lambda: defaultdict(dict))
slbVaDict = defaultdict(lambda: defaultdict(dict))
characters_to_remove = "^.*"
list_to_str_separator = "\r\n"
pattern_replace_port = r'(\bport\b)\s+(\d+)\s+(\w+)'
replacement_port = r'\1_\2_\3'


def iter_configfile(_config_file):
    while True:
        try:
            yield next(_config_file)
        except StopIteration as e:
            yield "Stop"


def process_string(input_string, input_characters_to_remove=characters_to_remove):
    # 删除指定字符
    processed_string = ''.join([char for char in input_string if char not in characters_to_remove])
    processed_string = processed_string.rstrip()
    # 将空格转换为下划线
    processed_string = processed_string.replace(' ', '_')
    return processed_string


def decode_unicode(text):
    # 以"%u"为分隔符分割字符串，然后将每一部分解码成16进制，再转换为Unicode字符
    text_list = []
    split_text = re.split(pattern_hex_split_unicode, text)
    split_text = [x for x in split_text if x]
    for text in split_text:
        if text.startswith('%u'):
            text_without_u = text.replace("%u", "")
            text_list.append(''.join(chr(int(text_without_u, 16))))
        else:
            text_list.append(text)
    text = ''.join(text_list)
    return text
 

def dict_to_excel(_dict):
    # 获取一级键
    first_level_keys = list(_dict.keys())
    # print("一级键:", first_level_keys)

    for sheet_name in first_level_keys:
        workbook.create_sheet(title=sheet_name)

    for first_key in first_level_keys:
        # 获取二级键
        second_level_keys = list(_dict[first_key].keys())
        # print("二级键:", second_level_keys)
        # 获取三级键
        # third_level_keys = list(_dict[first_level_keys[0]][second_level_keys[1]].keys())
        third_level_keys = []
        unique_third_level_keys = []
        for second_key in second_level_keys:
            for third_key in list(_dict[first_key][second_key].keys()):
                third_level_keys.append(third_key)
        for item in third_level_keys:
            if item not in unique_third_level_keys:
                unique_third_level_keys.append(item)
        # print("三级键:", unique_third_level_keys)
        function_name = first_key
        sheet = workbook[function_name]
        for index, value in enumerate(unique_third_level_keys, 2):
            # 从B1 单元格开始写行
            sheet.cell(row=1, column=index, value=value)
        for index, value in enumerate(second_level_keys, 2):
            # 将数据写入A列
            sheet.cell(row=index, column=1, value=value)
        row_index_begin, column_index_begin = 2, 2
        row_length = len(second_level_keys) + 1
        column_length = len(unique_third_level_keys) + 1
        for row_index in range(row_index_begin, row_length + 1):
            # 行
            for column_index in range(column_index_begin, column_length + 1):
                # 列
                row_index_data = sheet.cell(row=row_index, column=1).value
                column_index_data = sheet.cell(row=1, column=column_index).value
                try:
                    cell_value = _dict[function_name][row_index_data][column_index_data]
                    if isinstance(cell_value, str):
                        if re.match(pattern_hex_unicode, cell_value):
                            cell_value = decode_unicode(cell_value)
                    # print(_dict[function_name][row_index_data][column_index_data])
                    sheet.cell(row_index, column_index, value=cell_value)
                except KeyError:
                    sheet.cell(row_index, column_index, value="N/A")


def slbva_dict_to_excel(_dict, _sheet_name):
    _row_index_sequence = 2
    _unique_third_level_keys = []

    # 获取一级键
    _first_level_keys = list(_dict.keys())

    for _first_key in _first_level_keys:
        _third_level_keys = []
        # 获取二级键
        _second_level_keys = list(_dict[_first_key].keys())
        # print("二级键:", _second_level_keys)
        # 获取三级键
        for _second_key in _second_level_keys:
            try:
                for _third_key in list(_dict[_first_key][_second_key].keys()):
                    if _third_key not in _unique_third_level_keys:
                        _unique_third_level_keys.append(_third_key)
            except AttributeError:
                pass
        if "other" in _second_level_keys and "other" not in _unique_third_level_keys:
            _unique_third_level_keys.append("other")

    workbook.create_sheet(title=_sheet_name)
    sheet = workbook[_sheet_name]
    # 从B3 单元格开始写行
    for index, _value in enumerate(_unique_third_level_keys, 3):
        sheet.cell(row=1, column=index, value=_value)
    # 将数据写入A B列
    for _first_key in _first_level_keys:
        _second_level_keys = list(_dict[_first_key].keys())
        for index, _value in enumerate(_second_level_keys, _row_index_sequence):
            if _value != "other":
                sheet.cell(row=index, column=2, value=_value)
                sheet.cell(row=index, column=1, value=_first_key)
                _row_index_sequence += 1
    # 2 = 表头第一行 + range 不包含的上界；列同理
    _row_index_begin, _column_index_begin = 2, 3
    _row_index_sequence = 0
    _column_length = len(_unique_third_level_keys) + 1
    for _first_key in _first_level_keys:
        _second_level_keys = list(_dict[_first_key].keys())
        # -1 是去掉二级键 other
        _row_length = len(_second_level_keys) - 1
        for _first_row_index in range(_row_index_begin + _row_index_sequence, _row_length + _row_index_sequence + _row_index_begin):
            # 行
            for _column_index in range(_column_index_begin, _column_length + 1):
                # 列
                _first_row_index_data = sheet.cell(row=_first_row_index, column=1).value
                _second_row_index_data = sheet.cell(row=_first_row_index, column=2).value
                _column_index_data = sheet.cell(row=1, column=_column_index).value
                # print("##", _first_row_index_data, _second_row_index_data, _column_index_data)
                try:
                    _cell_value = _dict[_first_row_index_data][_second_row_index_data][_column_index_data]
                    if isinstance(_cell_value, str):
                        if re.match(pattern_hex_unicode, _cell_value):
                            _cell_value = decode_unicode(_cell_value)
                    sheet.cell(_first_row_index, _column_index, value=_cell_value)
                except KeyError:
                    sheet.cell(_first_row_index, _column_index, value="N/A")
                if _column_index == _column_length:
                    _other_value = _dict[_first_row_index_data]["other"]
                    if len(_other_value) > 1:
                        _cell_value = "\n".join(_other_value)
                    else:
                        _cell_value = "".join(_other_value)
                    sheet.cell(_first_row_index, _column_index + 1, value=_cell_value)
        _row_index_sequence += _row_length


configFile = open(file_path, 'r')
configFileIter = iter_configfile(configFile)
line_content = next(configFileIter)

while line_content != "Stop":
    for pattern in pattern_list:
        matches = re.match(pattern, line_content)
        if matches:
            blockConfig = []
            matchedLine = True
            masterKey = process_string(pattern.pattern)
            matchesConfigList = matches.group().split()
            line_content = next(configFileIter)
            # 跳过匹配行后面的第一个{
            if 'ip_nat' in masterKey:
                secondaryKey = matchesConfigList[3]
                line_content = next(configFileIter)
                while True:
                    endLineSign = re.match('^}$', line_content)
                    if endLineSign:
                        for config in blockConfig:
                            original_string = str(config)
                            string_list = original_string.split()
                            _member_value = ' '.join(str(item) for item in string_list[1:])
                            outputDict[masterKey][secondaryKey][string_list[0]] = _member_value
                        break
                    else:
                        blockConfig.append(str(line_content).strip())
                        line_content = next(configFileIter)
            elif 'health_check' in masterKey:
                secondaryKey = matchesConfigList[2]
                line_content = next(configFileIter)
                while True:
                    endLineSign = re.match('^}$', line_content)
                    if endLineSign:
                        if matchesConfigList[3] == "interval":
                            _member_value = matchesConfigList[4]
                            outputDict[masterKey][secondaryKey]['interval'] = _member_value
                        if matchesConfigList[5] == "retry":
                            _member_value = matchesConfigList[6]
                            outputDict[masterKey][secondaryKey]['retry'] = _member_value
                        if matchesConfigList[7] == "timeout":
                            _member_value = matchesConfigList[8]
                            outputDict[masterKey][secondaryKey]['timeout'] = _member_value
                        if matchesConfigList[9] == "up-check-cnt":
                            _member_value = matchesConfigList[10]
                            outputDict[masterKey][secondaryKey]['up-check-cnt'] = _member_value
                        for config in blockConfig:
                            original_string = str(config)
                            string_list = original_string.split()
                            if string_list[0] == "wait-all-retry":
                                _member_value = True
                                outputDict[masterKey][secondaryKey][string_list[0]] = _member_value
                            else:
                                _member_value = ' '.join(str(item) for item in string_list[1:])
                                outputDict[masterKey][secondaryKey][string_list[0]] = _member_value

                        break
                    else:
                        blockConfig.append(str(line_content).strip())
                        line_content = next(configFileIter)
            elif 'slb_pool' in masterKey:
                _pool_member = []
                secondaryKey = matchesConfigList[2]
                line_content = next(configFileIter)
                while True:
                    endLineSign = re.match('^}$', line_content)
                    if endLineSign:
                        if matchesConfigList[3] != '':
                            outputDict[masterKey][secondaryKey]["procotol"] = matchesConfigList[3]
                        for config in blockConfig:
                            original_string = str(config)
                            config_string_list = original_string.split()
                            if config_string_list[0] == 'member':
                                _pool_member.append(config_string_list[1])
                            else:
                                _member_value = ' '.join(str(item) for item in config_string_list[1:])
                                outputDict[masterKey][secondaryKey][config_string_list[0]] = _member_value
                        list_to_str = list_to_str_separator.join(map(str, _pool_member))
                        outputDict[masterKey][secondaryKey]["member"] = list_to_str
                        break
                    else:
                        blockConfig.append(str(line_content).strip())
                        line_content = next(configFileIter)
            elif 'slb_virtual-address' in masterKey:
                otherConfig = []
                vaKey = matchesConfigList[2]
                line_content = next(configFileIter)
                while True:
                    endLineSign = re.match('^}$', line_content)
                    if endLineSign:
                        if matchesConfigList[3] != '':
                            pass
                        for config in blockConfig:
                            if config.startswith("port"):
                                current_key = str(config)
                                port_config_list = []
                            elif config.startswith("{"):
                                port_config_begin = True
                            elif config.startswith("}") and port_config_begin:
                                list_to_dict = {}
                                for item in port_config_list:
                                    parts = item.split(' ', 1)  # 分割字符串为最多两部分，以便将第一个单词作为键
                                    if len(parts) > 1:
                                        key = parts[0]
                                        value = parts[1]
                                        list_to_dict[key] = value
                                    else:
                                        list_to_dict[item] = ''  # 对于只有一个单词的情况，值为空字符串
                                slbVaDict[vaKey][current_key] = list_to_dict
                                del port_config_begin, current_key
                            elif 'port_config_begin' in locals() and 'current_key' in locals():
                                port_config_list.append(config)
                            else:
                                otherConfig.append(config)
                        slbVaDict[vaKey]["other"] = otherConfig
                        break
                    else:
                        blockConfig.append(str(line_content).strip())
                        line_content = next(configFileIter)
            if r'slb_profile\b(?!(?:node|persist))$' in masterKey:
                masterKey = "slb_profile"
                secondaryKey = matchesConfigList[3]
                line_content = next(configFileIter)
                while True:
                    endLineSign = re.match('^}$', line_content)
                    if endLineSign:
                        for config in blockConfig:
                            original_string = str(config)
                            config_string_list = original_string.split()
                            if len(config_string_list) == 1:
                                _value = True
                                outputDict[masterKey][secondaryKey][config_string_list[0]] = _value
                            else:
                                _value = ' '.join(str(item) for item in config_string_list[1:])
                                outputDict[masterKey][secondaryKey][config_string_list[0]] = _value
                        break
                    else:
                        blockConfig.append(str(line_content).strip())
                        line_content = next(configFileIter)
            if 'slb_profile_persist' in masterKey:
                secondaryKey = matchesConfigList[4]
                line_content = next(configFileIter)
                while True:
                    endLineSign = re.match('^}$', line_content)
                    if endLineSign:
                        if len(blockConfig) == 0:
                            outputDict[masterKey][secondaryKey]['Null'] = True
                            break
                        else:
                            for config in blockConfig:
                                original_string = str(config)
                                string_list = original_string.split()
                                _member_value = ' '.join(str(item) for item in string_list[1:])
                                if _member_value == '':
                                    _member_value = True
                                outputDict[masterKey][secondaryKey][string_list[0]] = _member_value
                            break
                    else:
                        blockConfig.append(str(line_content).strip())
                        line_content = next(configFileIter)
                # pass
    line_content = next(configFileIter)
else:
    if os.path.exists("output.xlsx"):
        os.remove("output.xlsx")
    workbook = openpyxl.Workbook()
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    dict_to_excel(outputDict)
    slbva_dict_to_excel(slbVaDict, "slb_virtual-address")
    workbook.save("output.xlsx")

